from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference

# Загрузка файла
wb = load_workbook(r"./Forms_1-14.xlsx")
ws = wb['Лист1']
labels = ["Да", "Нет"]

for col in range(2, 14):  # Столбцы B (2) до AV (49)
    chart = PieChart()
    # Данные из строк 3–4
    data = Reference(ws, min_col=col, min_row=3, max_row=4)
    # Добавляем данные в диаграмму
    chart.add_data(data, titles_from_data=False)
    # Создаем фиктивные подписи через Reference
    # Для этого создадим временный диапазон в Excel с фиксированными подписями
    for i, label in enumerate(labels, start=3):
        ws.cell(row=i, column=15).value = label  # Записываем подписи в колонку 50 (столбец AX)
    # Создаем Reference для подписей
    labels_ref = Reference(ws, min_col=15, min_row=3, max_row=4)
    # Устанавливаем подписи
    chart.set_categories(labels_ref)
    # Заголовок диаграммы (вопрос из первой строки)
    chart.title = ws.cell(row=1, column=col).value
    # Размещаем диаграмму на листе
    ws.add_chart(chart, f"A{col * 5}")
# Сохранение файла
wb.save(r"./АКАДОС_с_диаграммами_1-14.xlsx")
