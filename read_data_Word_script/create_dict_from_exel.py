from openpyxl import load_workbook

book = load_workbook('холл — копия.xlsx')   # file name
list_1 = book.get_sheet_by_name('Лист1')    # sheet name

guests = {}     # dictionary

# cycle for appending elements to dictionary, without None elements
for i in range(5, 496 + 1):
    if (i == 467 or
            i == 460 or
            i == 439 or
            i == 338 or
            i == 173):
        continue
    else:
        guests[list_1.cell(row=i, column=2).value] = []

# test print
# for key in guests:
#     print(key, guests[key])
